import os
import pandas as pd
from openpyxl import Workbook

# Базовая папка с результатами
base_folder = "results/configuration_2/"
base_folder = "results/configuration_2/"

# Создаем Excel-файл
output_file = base_folder + "/experiment_results.xlsx"

# Создаем новый Workbook
wb = Workbook()
ws = wb.active

# Записываем название папки configuration_2 в ячейку A1
ws.cell(row=1, column=1, value=base_folder)

# Начальная колонка для размещения данных
start_columns = [1, 8, 15, 22, 29]  # Колонки A, H, N, T (2, 8, 14, 20 в индексации openpyxl)

# Запись текстовых меток в ячейки A2, G2, M2, S2
labels = ["RR", "WRRs", "WRR", "LC", "WLC"]
label_columns = [1, 8, 15, 22, 29]  # Колонки A, G, M, S (1, 7, 13, 19 в индексации openpyxl)
for col_idx, label in zip(label_columns, labels):
    ws.cell(row=2, column=col_idx, value=label)

# Чтение всех CSV-файлов в базовой папке
csv_files = [f for f in os.listdir(base_folder) if f.endswith('.csv')]
if not csv_files:
    print(f"В папке '{base_folder}' нет CSV-файлов.")
else:
    # Разделяем файлы по меткам (RR, WRR, LC, WLC)
    file_groups = {label: [] for label in labels}
    for csv_file in csv_files:
        for label in labels:
            # Проверяем точное соответствие метки началу имени файла
            if csv_file.startswith(label):
                file_groups[label].append(csv_file)
                break  # Переходим к следующему файлу после нахождения соответствия

    # Обработка файлов для каждой метки
    for label, start_col in zip(labels, start_columns):
        if not file_groups[label]:
            print(f"Нет CSV-файлов для метки '{label}'.")
            continue

        # Берем первый CSV-файл из группы
        csv_file = os.path.join(base_folder, file_groups[label][0])
        df = pd.read_csv(csv_file)

        # Запись данных в Excel (начинаем с строки 4)
        for row_idx, row in enumerate(df.values, start=4):  # Сдвиг на две строки вниз
            for col_idx, value in enumerate(row, start=start_col):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Запись заголовков (сдвиг на две строки вниз)
        for col_idx, header in enumerate(df.columns, start=start_col):
            ws.cell(row=3, column=col_idx, value=header)  # Заголовки пишутся в строке 3

# Сохраняем Excel-файл
wb.save(output_file)
print(f"Файл '{output_file}' успешно создан.")