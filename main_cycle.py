from node import Server
from distributor import RoundRobin, WeightedRoundRobin, WeightedRoundRobinStatic, LeastConnection, WeightedLeastConnection
import random
import csv
from typing import List, Dict
import os
import pandas as pd
from openpyxl import Workbook


def transfer_data_from_csv_to_exel(config):
    # Базовая папка с результатами
    base_folder = f"results/configuration_{config}/"

    # Создаем Excel-файл
    output_file = base_folder + "/experiment_results.xlsx"

    # Создаем новый Workbook
    wb = Workbook()
    ws = wb.active

    # Записываем название папки configuration_2 в ячейку A1
    ws.cell(row=1, column=1, value=base_folder)

    # Начальная колонка для размещения данных
    start_columns = [1, 8, 15, 22, 29]  # Колонки A, H, N, T (2, 8, 14, 20 в индексации openpyxl)

    # Запись текстовых меток в ячейки A2, G2, M2, S2
    labels = ["RR", "WRRs", "WRR", "LC", "WLC"]
    label_columns = [1, 8, 15, 22, 29]  # Колонки A, G, M, S (1, 7, 13, 19 в индексации openpyxl)
    for col_idx, label in zip(label_columns, labels):
        ws.cell(row=2, column=col_idx, value=label)

    # Чтение всех CSV-файлов в базовой папке
    csv_files = [f for f in os.listdir(base_folder) if f.endswith('.csv')]
    if not csv_files:
        print(f"В папке '{base_folder}' нет CSV-файлов.")
    else:
        # Разделяем файлы по меткам (RR, WRR, LC, WLC)
        file_groups = {label: [] for label in labels}
        for csv_file in csv_files:
            for label in labels:
                # Проверяем точное соответствие метки началу имени файла
                if csv_file.startswith(label):
                    file_groups[label].append(csv_file)
                    break  # Переходим к следующему файлу после нахождения соответствия

        # Обработка файлов для каждой метки
        for label, start_col in zip(labels, start_columns):
            if not file_groups[label]:
                print(f"Нет CSV-файлов для метки '{label}'.")
                continue

            # Берем первый CSV-файл из группы
            csv_file = os.path.join(base_folder, file_groups[label][0])
            df = pd.read_csv(csv_file)

            # Запись данных в Excel (начинаем с строки 4)
            for row_idx, row in enumerate(df.values, start=4):  # Сдвиг на две строки вниз
                for col_idx, value in enumerate(row, start=start_col):
                    ws.cell(row=row_idx, column=col_idx, value=value)

            # Запись заголовков (сдвиг на две строки вниз)
            for col_idx, header in enumerate(df.columns, start=start_col):
                ws.cell(row=3, column=col_idx, value=header)  # Заголовки пишутся в строке 3

    # Сохраняем Excel-файл
    wb.save(output_file)
    print(f"Файл '{output_file}' успешно создан.")


def save_servers_to_csv(servers, filename: str):
    """
    Сохраняет данные серверов в CSV файл

    :param servers: Список словарей с данными серверов
    :param filename: Имя файла для сохранения
    """
    headers = [
        'Node',
        'Power Load (%)',
        'Network Load (%)',
        'Tasks Load (pieces)',
        'Work time (sec)',
        'Total Calculated Tasks'
    ]

    with open(filename, mode='w', newline='') as file:
        writer = csv.DictWriter(file, fieldnames=headers)
        writer.writeheader()

        for server in servers:
            # Рассчитываем средневзвешенные значения
            weighted_load = sum(server.cpu_load_history) / len(server.cpu_load_history) if server.cpu_load_history else 0
            weighted_network = sum(server.network_load_history) / len(server.network_load_history) if server.network_load_history else 0
            weighted_tasks = sum(server.tasks_history) / len(server.tasks_history) if server.tasks_history else 0

            writer.writerow({
                'Node': server.server_id,
                'Power Load (%)': round(weighted_load, 4),
                'Network Load (%)': round(weighted_network, 4),
                'Tasks Load (pieces)': round(weighted_tasks, 4),
                'Work time (sec)': round(server.total_work_time),
                'Total Calculated Tasks': server.processed_tasks
            })



def simulate_cpu_load(task_duration, tasks_per_second, simulation_time):
    """
    Рассчитывает нагрузку CPU с учетом отброшенных задач.

    :param task_duration: время выполнения одной задачи в секундах
    :param tasks_per_second: среднее количество задач в секунду
    :param simulation_time: общее время симуляции в секундах
    """
    print(f"\nСимуляция нагрузки CPU ({simulation_time} сек)")
    print(f"Параметры: {task_duration=} сек, {tasks_per_second=} задач/сек")
    print("-" * 50)

    stats = []
    total_dropped = 0  # Общее количество отброшенных задач

    for second in range(int(simulation_time)):
        # Максимальное количество задач, которые можно выполнить за 1 секунду
        max_tasks = int(1.0 / task_duration) if task_duration > 0 else float('inf')

        # Фактическое количество задач (может быть больше, чем CPU может обработать)
        incoming_tasks = tasks_per_second

        # Количество выполненных и отброшенных задач
        processed_tasks = min(incoming_tasks, max_tasks)
        dropped_tasks = max(0, incoming_tasks - max_tasks)
        total_dropped += dropped_tasks

        # Нагрузка CPU (всегда <= 100%)
        cpu_load = min(100.0, processed_tasks * task_duration * 100)

        stats.append((second, cpu_load, processed_tasks, dropped_tasks))

    # Вывод результатов
    print("\nРезультаты:")
    print(f"{'Секунда':<10}{'Нагрузка':<10}{'Задач выполнено':<15}{'Отброшено':<10}")
    for sec, load, processed, dropped in stats:
        print(f"{sec:<10}{load:<10.1f}%{processed:<15}{dropped:<10}")

    print(f"\nВсего отброшено задач: {total_dropped}")


# configuration_1

configurations = {
    1:
        [Server(server_id=1, bu_power=1.22, bandwidth_bytes=80_000),
        Server(server_id=2, bu_power=1.22, bandwidth_bytes=80_000),
        Server(server_id=3, bu_power=1.22, bandwidth_bytes=80_000),
        Server(server_id=4, bu_power=1.22, bandwidth_bytes=80_000),
        Server(server_id=5, bu_power=1.22, bandwidth_bytes=80_000),
        Server(server_id=6, bu_power=1.22, bandwidth_bytes=80_000),
        Server(server_id=7, bu_power=1.22, bandwidth_bytes=80_000),
        Server(server_id=8, bu_power=1.22, bandwidth_bytes=80_000),
        Server(server_id=9, bu_power=1.22, bandwidth_bytes=80_000),
        Server(server_id=10, bu_power=1.22, bandwidth_bytes=80_000),
        Server(server_id=11, bu_power=1.22, bandwidth_bytes=80_000),
        Server(server_id=12, bu_power=1.22, bandwidth_bytes=80_000),],
    2:
        [Server(server_id=1, bu_power=1, bandwidth_bytes=80_000),
        Server(server_id=2, bu_power=1, bandwidth_bytes=80_000),
        Server(server_id=3, bu_power=1, bandwidth_bytes=80_000),
        Server(server_id=4, bu_power=1, bandwidth_bytes=80_000),
        Server(server_id=5, bu_power=1.22, bandwidth_bytes=80_000),
        Server(server_id=6, bu_power=1.22, bandwidth_bytes=80_000),
        Server(server_id=7, bu_power=1.22, bandwidth_bytes=80_000),
        Server(server_id=8, bu_power=1.22, bandwidth_bytes=80_000),
        Server(server_id=9, bu_power=1.22, bandwidth_bytes=80_000),
        Server(server_id=10, bu_power=1.22, bandwidth_bytes=80_000),
        Server(server_id=11, bu_power=1.22, bandwidth_bytes=80_000),
        Server(server_id=12, bu_power=2.2, bandwidth_bytes=80_000),],
    3:
        [Server(server_id=1, bu_power=1, bandwidth_bytes=80_000),
        Server(server_id=2, bu_power=1, bandwidth_bytes=80_000),
        Server(server_id=3, bu_power=1, bandwidth_bytes=80_000),
        Server(server_id=4, bu_power=1, bandwidth_bytes=80_000),
        Server(server_id=5, bu_power=1, bandwidth_bytes=80_000),
        Server(server_id=6, bu_power=1, bandwidth_bytes=80_000),
        Server(server_id=7, bu_power=1, bandwidth_bytes=80_000),
        Server(server_id=8, bu_power=1.22, bandwidth_bytes=80_000),
        Server(server_id=9, bu_power=1.22, bandwidth_bytes=80_000),
        Server(server_id=10, bu_power=1.22, bandwidth_bytes=80_000),
        Server(server_id=11, bu_power=1.22, bandwidth_bytes=80_000),
        Server(server_id=12, bu_power=2.2, bandwidth_bytes=80_000),],
    4:
        [Server(server_id=1, bu_power=1, bandwidth_bytes=80_000),
        Server(server_id=2, bu_power=1.22, bandwidth_bytes=80_000),
        Server(server_id=3, bu_power=1.22, bandwidth_bytes=80_000),
        Server(server_id=4, bu_power=1.22, bandwidth_bytes=80_000),
        Server(server_id=5, bu_power=1.22, bandwidth_bytes=80_000),
        Server(server_id=6, bu_power=2.2, bandwidth_bytes=80_000),
        Server(server_id=7, bu_power=2.2, bandwidth_bytes=80_000),
        Server(server_id=8, bu_power=2.2, bandwidth_bytes=80_000),
        Server(server_id=9, bu_power=2.2, bandwidth_bytes=80_000),
        Server(server_id=10, bu_power=2.2, bandwidth_bytes=80_000),
        Server(server_id=11, bu_power=2.2, bandwidth_bytes=80_000),
        Server(server_id=12, bu_power=2.2, bandwidth_bytes=80_000),],
}


def generate_repeating_task_list(total_tasks):
    # Базовый паттерн
    pattern = [0.02] * 6 + [0.1] * 3 + [0.28] * 1

    # Вычисляем количество полных паттернов и остаток
    full_patterns = total_tasks // len(pattern)
    remainder = total_tasks % len(pattern)

    # Формируем итоговый список
    task_list = pattern * full_patterns + pattern[:remainder]

    return task_list


# Пример использования
if __name__ == "__main__":

    config = 4
    folder_path = f"results/configuration_{config}/"
    servers = configurations[config] #[::-1]
    #random.shuffle(servers)

    # Параметры симуляции
    task_time = 0.02 # каждая задача выполняется 0.3 секунды
    task_size = 500

    tasks_config_experiment_3 = {1: 500, 2: 500, 3: 500, 4: 500}
    tasks_per_second = tasks_config_experiment_3[config]
    tasks = generate_repeating_task_list(tasks_per_second)
    print(tasks_per_second)
    print(tasks)

    # tasks_per_second = 184  # 5 задач в секунду
    simulation_time = 120  # симулируем 10 секунд

    # task_time, task_size, tasks_per_second
    # 6 : 3 : 1
    # 0.02 0.1 0.28

    #tasks = [[0.02, 500, 110], [0.1, 500, 55], [0.28, 500, 18]]
    #tasks = [[0.02, 500, 110], [0.1, 500, 55], [0.28, 500, 18]]
    #tasks = [[0.02, 500, 105], [0.1, 500, 53], [0.28, 500, 17]]
    #tasks = [[0.02, 500, 159], [0.1, 500, 80], [0.28, 500, 26]]


    distributor = WeightedRoundRobin(servers)
    distributors = [RoundRobin(servers),
                    WeightedRoundRobin(servers),
                    WeightedRoundRobinStatic(servers),
                    LeastConnection(servers),
                    WeightedLeastConnection(servers)]

    for distributor in distributors:
        for server in servers:
            server.reset_for_new_second()
        task_num = 1
        for i in range(1, simulation_time+1):

            # for task_type in tasks:
            #     task_num += 1
            #     task_time, task_size, tasks_per_second = task_type
            #     for _ in range(tasks_per_second):
            #         distributor.distribute_task(task_time, task_size)

            # for task in range(tasks_per_second):
            #     distributor.distribute_task(task_time, task_size)


            # специально для 3 эксперимента такой обход:
            for task in tasks:
                distributor.distribute_task(task_time, task_size)


            if type(distributor).__name__ == "LeastConnection" and (i % 2 == 0):
                # потому что если обновлться каждое изменение подключений, то распредлеление будет идти косо
                distributor.updated_nodes_connections(servers)



            for server in servers:
                server.reset_for_new_second()

        for server in servers:
            server.cpu_load_history.pop()
            server.network_load_history.pop()
            server.tasks_history.pop()


        for server in servers:
            print(f"Server_{server.server_id}")
            #print(f"{'Секунда':<10}{'Нагрузка':<10}{'Нагрузка сети':<10}{'Задач решено':<5}")
            # for i in range(len(server.cpu_load_history)-1):
            #     load = round(server.cpu_load_history[i], 4)
            #     network_load = round(server.network_load_history[i], 4)
            #     tasks = server.tasks_history[i]
            #
            #     print(f"{i+1:<10}{load:<10.1f}{network_load:<16.1f}{tasks:<5}")

            print("-"*60)
            print(f"Решено задач: {server.processed_tasks}")
            print(f"Отклонено задач конкретно этим сервером: {server.dropped_tasks}")
            print(f"Средняя загрузка сервера: {sum(server.cpu_load_history) / len(server.cpu_load_history)}")
            print(f"Средняя загрузка сети сервера: {sum(server.network_load_history) / len(server.network_load_history)}")
            print(f"Время работы: {server.total_work_time}")
            print("-" * 60)

        total_calculated = sum(server.processed_tasks for server in servers)
        total_rejected = sum(server.dropped_tasks for server in servers)
        total_generated_tasks = tasks_per_second * simulation_time
        print(f"Всего создано задач: {total_generated_tasks}\n"
              f"Всего решено: {total_calculated}\n"
              f"Всего отклонено: {distributor.rejected_tasks}")


        csv_names = {"RoundRobin": "RR.csv",
                     "WeightedRoundRobinStatic": "WRRs.csv",
                     "WeightedRoundRobin": "WRR.csv",
                     "LeastConnection": "LC.csv",
                     "WeightedLeastConnection": "WLC.csv"}
        res_csv_name = csv_names[type(distributor).__name__]
        save_servers_to_csv(servers, folder_path + res_csv_name)
        #print(servers[-1].cpu_load_history)


        servers_load = []
        for server in servers:
            weighted_load = sum(server.cpu_load_history) / len(server.cpu_load_history) if server.cpu_load_history else 0
            weighted_network = sum(server.network_load_history) / len(server.network_load_history) if server.network_load_history else 0
            weighted_tasks = sum(server.tasks_history) / len(server.tasks_history) if server.tasks_history else 0
            servers_load.append(weighted_load)


        def calculate_std_dev(data):
            if len(data) == 0:
                return 0  # Возвращаем 0 для пустого списка

            mean = sum(data) / len(data)  # Вычисляем среднее
            variance = sum((x - mean) ** 2 for x in data) / len(data)  # Вычисляем дисперсию
            std_dev = variance ** 0.5  # Стандартное отклонение - это корень из дисперсии
            return std_dev

        print("Servers load from 1 to 12: ", servers_load)
        import numpy as np
        std_dev = np.std(servers_load)
        print(f"Стандартное отклонение: {std_dev}")
        print(f"Стандартное отклонение: {calculate_std_dev(servers_load)}")

        for server in servers:
            server.reset()

    transfer_data_from_csv_to_exel(config)